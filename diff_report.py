import os
import json
import warnings
import urllib.parse
import zipfile
import xml.etree.ElementTree as ET
import re
from datetime import datetime
import openpyxl

warnings.simplefilter("ignore", UserWarning)

DEFAULT_LOG_FOLDER = r"D:\Pzone\excel_compare"
DEFAULT_DIFF_REPORT_DIR = None

def get_excel_metadata(file_path):
    file_path = file_path.strip('"')
    if not os.path.exists(file_path):
        return "Unknown", "Unknown"
    
    mod_timestamp = os.path.getmtime(file_path)
    last_modified_date = datetime.fromtimestamp(mod_timestamp).strftime('%Y-%m-%d %H:%M:%S')
    
    author = "Unknown"
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if hasattr(wb.properties, 'lastModifiedBy') and wb.properties.lastModifiedBy:
            author = wb.properties.lastModifiedBy
        elif wb.properties.creator:
            author = wb.properties.creator
        wb.close()
    except Exception:
        author = "Unable to read"
        
    return author, last_modified_date

def clean_target_path(raw_target):
    if not raw_target:
        return ""
    cleaned = urllib.parse.unquote(raw_target)
    for prefix in ['file:///', 'file://', 'file:/']:
        if cleaned.startswith(prefix):
            cleaned = cleaned[len(prefix):]
            break
    if re.match(r'^/[A-Za-z]:', cleaned):
        cleaned = cleaned[1:]
    cleaned = cleaned.replace('/', '\\')
    return cleaned.strip()

def extract_external_refs(file_path):
    ref_map = {}
    file_path = file_path.strip('"')
    
    if not os.path.exists(file_path):
        return ref_map

    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            namelist = z.namelist()
            rels_path = 'xl/_rels/workbook.xml.rels'
            ext_target_map = {} 
            
            if rels_path in namelist:
                rels_content = z.read(rels_path)
                root = ET.fromstring(rels_content)
                ns = {'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                
                for rel in root.findall('rels:Relationship', ns):
                    target = rel.get('Target', '')
                    type_str = rel.get('Type', '')
                    if 'externalLink' in type_str or 'externalLinks' in target:
                        match_id = re.search(r'externalLink(\d+)\.xml', target)
                        if match_id:
                            ext_target_map[int(match_id.group(1))] = target

            for idx, link_target in ext_target_map.items():
                link_target_clean = link_target.lstrip('/')
                dir_name = os.path.dirname(link_target_clean)
                base_name = os.path.basename(link_target_clean)
                rels_file_path = os.path.join('xl', dir_name, '_rels', f"{base_name}.rels").replace('\\', '/')
                
                real_file_path = ""
                if rels_file_path in namelist:
                    try:
                        sub_rels_content = z.read(rels_file_path)
                        sub_root = ET.fromstring(sub_rels_content)
                        sub_ns = {'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                        for sub_rel in sub_root.findall('rels:Relationship', sub_ns):
                            target = sub_rel.get('Target', '')
                            if target:
                                real_file_path = clean_target_path(target)
                                break
                    except Exception:
                        pass
                
                if real_file_path:
                    ref_map[idx] = real_file_path
    except Exception:
        pass
        
    return ref_map

def pretty_formula(formula, ref_map):
    if not formula:
        return ""
    formula = str(formula).strip()
    while formula.startswith('=='):
        formula = '=' + formula[2:]
    if not formula.startswith('='):
        formula = '=' + formula

    formula = re.sub(r"=''\[(\d+)\]([^!]+)''!", r"[\1]\2!", formula)

    def replace_with_sheet(match):
        num = int(match.group(1))
        sheet_name = match.group(2).strip("'\"")
        if num in ref_map:
            full_path = ref_map[num]
            directory, filename = os.path.split(full_path)
            if directory and filename:
                return f"'{directory}\\[{filename}]{sheet_name}'!"
            return f"'{full_path}'!{sheet_name}!"
        return match.group(0)
    
    formula = re.sub(r'\[(\d+)\]([^!]+)\!', replace_with_sheet, formula)
    
    def replace_general(match):
        num = int(match.group(1))
        if num in ref_map:
            full_path = ref_map[num]
            directory, filename = os.path.split(full_path)
            if directory and filename:
                return f"'{directory}\\[{filename}]'"
            return f"'{full_path}'"
        return match.group(0)
        
    formula = re.sub(r'\[(\d+)\]', replace_general, formula)
    formula = formula.replace("=''", "='").replace("''!", "'!").replace("''", "'")
    return formula

def excel_to_dict(file_path):
    file_path = file_path.strip('"')
    ref_map = extract_external_refs(file_path)

    wb_val = openpyxl.load_workbook(file_path, data_only=True)
    wb_form = openpyxl.load_workbook(file_path, data_only=False)
    
    data_dict = {}
    for sheet_name in wb_val.sheetnames:
        sheet_v = wb_val[sheet_name]
        sheet_f = wb_form[sheet_name]
        sheet_data = {}
        
        for row_idx, row in enumerate(sheet_v.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is not None:
                    coordinate = cell.coordinate
                    val = cell.value
                    
                    form_cell = sheet_f.cell(row=row_idx, column=col_idx)
                    formula = ""
                    if form_cell.value is not None and str(form_cell.value).startswith('='):
                        formula = pretty_formula(str(form_cell.value), ref_map)
                    
                    sheet_data[coordinate] = {
                        "value": val,
                        "formula": formula
                    }
        data_dict[sheet_name] = sheet_data
        
    wb_val.close()
    wb_form.close()
    return data_dict

def generate_diff_report(old_data, new_data, old_file_path, new_file_path, output_dir=DEFAULT_DIFF_REPORT_DIR, include_unchanged_cells=False):
    if output_dir is None:
        output_dir = os.path.join(DEFAULT_LOG_FOLDER, 'diff_reports')
    os.makedirs(output_dir, exist_ok=True)
    
    diff_data = prepare_diff_data(old_data, new_data, old_file_path, new_file_path)
    
    old_sheets = set(old_data.keys())
    new_sheets = set(new_data.keys())
    matched_sheets = sorted(list(old_sheets & new_sheets))
    old_only_sheets = sorted(list(old_sheets - new_sheets))
    new_only_sheets = sorted(list(new_sheets - old_sheets))
    
    sheet_meta = {
        "matched": matched_sheets,
        "oldOnly": old_only_sheets,
        "newOnly": new_only_sheets
    }
    
    html_content = generate_html_content(diff_data, sheet_meta, old_file_path, new_file_path, include_unchanged_cells)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    new_filename = os.path.basename(new_file_path.strip('"'))
    filename_base = os.path.splitext(new_filename)[0]
    filename = f"diff_matrix_report_{filename_base}_{timestamp}.html"
    output_path = os.path.join(output_dir, filename)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"[diff-report] Report generated: {output_path}")
    return output_path

def prepare_diff_data(old_data, new_data, old_file_path, new_file_path):
    old_sheets_set = set(old_data.keys())
    new_sheets_set = set(new_data.keys())
    all_sheets = old_sheets_set | new_sheets_set
    change_items = []
    
    for sheet_name in all_sheets:
        is_matched = (sheet_name in old_sheets_set and sheet_name in new_sheets_set)
        is_old_only = (sheet_name in old_sheets_set and sheet_name not in new_sheets_set)
        is_new_only = (sheet_name not in old_sheets_set and sheet_name in new_sheets_set)
        
        old_sheet = old_data.get(sheet_name, {})
        new_sheet = new_data.get(sheet_name, {})
        all_addresses = set(old_sheet.keys()) | set(new_sheet.keys())
        
        for address in all_addresses:
            old_cell = old_sheet.get(address, {})
            new_cell = new_sheet.get(address, {})
            
            old_val = old_cell.get('value', '')
            new_val = new_cell.get('value', '')
            old_formula = old_cell.get('formula', '')
            new_formula = new_cell.get('formula', '')
            
            val_diff_str = ""
            has_diff = False
            
            if is_old_only:
                has_diff = True
            elif is_new_only:
                has_diff = True
            else:
                has_diff = (old_cell != new_cell)
                
                try:
                    if str(old_val).strip() == "" and str(new_val).strip() != "":
                        val_diff_str = str(float(new_val))
                    elif str(old_val).strip() != "" and str(new_val).strip() == "":
                        diff_val = 0.0 - float(old_val)
                        val_diff_str = str(diff_val)
                    elif str(old_val) != "" and str(new_val) != "":
                        diff_val = float(new_val) - float(old_val)
                        if diff_val != 0:
                            val_diff_str = str(diff_val)
                except (ValueError, TypeError):
                    if old_val != new_val:
                        val_diff_str = f'<span class="diff-deleted">{old_val}</span> <strong>-></strong> <span class="diff-added">{new_val}</span>'
            
            change_items.append({
                'sheet': sheet_name,
                'address': address,
                'oldVal': str(old_val) if old_val is not None else '',
                'newVal': str(new_val) if new_val is not None else '',
                'valDiff': val_diff_str,
                'oldFormula': old_formula,
                'newFormula': new_formula,
                'hasDiff': has_diff,
                'isOldOnly': is_old_only,
                'isNewOnly': is_new_only
            })
                
    timestamp_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    events = [{
        "eventIndex": 1,
        "timestamp": timestamp_str,
        "changes": change_items
    }]
        
    file_entry = {
        "file": os.path.basename(new_file_path.strip('"')),
        "filePath": new_file_path.strip('"'),
        "events": events
    }
    return [file_entry]

def generate_html_content(diff_data, sheet_meta, old_file_path, new_file_path, include_unchanged_cells):
    json_data = json.dumps(diff_data, ensure_ascii=False).replace("</script>", "<\\/script>")
    sheet_meta_json = json.dumps(sheet_meta, ensure_ascii=False)
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    old_file_path = old_file_path.strip('"')
    new_file_path = new_file_path.strip('"')
    
    old_author, old_mod = get_excel_metadata(old_file_path)
    new_author, new_mod = get_excel_metadata(new_file_path)
    
    old_path_safe = urllib.parse.quote(old_file_path.replace('\\', '/'))
    new_path_safe = urllib.parse.quote(new_file_path.replace('\\', '/'))
    
    old_url = f"file:///{old_path_safe}"
    new_url = f"file:///{new_path_safe}"
    
    if include_unchanged_cells:
        unchanged_checkbox_html = """
                <div>
                    <label style="cursor: pointer; font-weight: bold; display: inline-flex; align-items: center; gap: 5px;">
                        <input type="checkbox" id="showUnchanged" onchange="onFilterChange()" checked> Show Unchanged & Unpaired
                    </label>
                </div>"""
        include_unchanged_js_flag = "true"
    else:
        unchanged_checkbox_html = """
                <div>
                    <label style="cursor: pointer; font-weight: bold; display: inline-flex; align-items: center; gap: 5px;">
                        <input type="checkbox" id="showUnchanged" onchange="onFilterChange()"> Show Unchanged & Unpaired
                    </label>
                </div>"""
        include_unchanged_js_flag = "false"
    
    html_template = r"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Excel Difference History Report</title>
    <style>
        :root {
            --bg-color: #000000;
            --text-color: #00ff00;
            --border-color: #00ff00;
            --header-bg: #001100;
            --btn-bg: #000000;
            --btn-hover-bg: #00ff00;
            --btn-hover-color: #000000;
            --added-bg: #003300;
            --added-color: #00ff00;
            --deleted-bg: #330000;
            --deleted-color: #ff6666;
            --link-color: #00ff00;
            --shadow: 0 0 8px rgba(0,255,0,0.2);
            --font-family: Consolas, "Courier New", Courier, monospace;
        }

        [data-theme="traditional"] {
            --bg-color: #f4f6f9;
            --text-color: #333333;
            --border-color: #cccccc;
            --header-bg: #e9ecef;
            --btn-bg: #ffffff;
            --btn-hover-bg: #0056b3;
            --btn-hover-color: #ffffff;
            --added-bg: #d4edda;
            --added-color: #155724;
            --deleted-bg: #f8d7da;
            --deleted-color: #721c24;
            --link-color: #0068c9;
            --shadow: 0 2px 4px rgba(0,0,0,0.05);
            --font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
        }

        body {
            background-color: var(--bg-color);
            color: var(--text-color);
            font-family: var(--font-family);
            line-height: 1.4;
            max-width: 98%;
            margin: 10px auto;
            padding: 0 10px;
            transition: background 0.3s, color 0.3s;
        }
        h1, h3, h4 { color: var(--text-color); margin-top: 0; }
        .file-info, .section-box {
            background-color: var(--bg-color);
            border: 1px solid var(--border-color);
            padding: 15px;
            margin-bottom: 15px;
            box-shadow: var(--shadow);
            border-radius: 5px;
        }
        .file-link {
            text-decoration: underline;
            color: var(--link-color);
            font-weight: bold;
        }
        .meta-text {
            font-size: 0.9em;
            opacity: 0.85;
            margin-top: 2px;
            margin-bottom: 10px;
        }
        .btn, select.btn {
            background-color: var(--btn-bg);
            color: var(--text-color);
            border: 1px solid var(--border-color);
            padding: 6px 12px;
            cursor: pointer;
            font-family: var(--font-family);
            font-weight: bold;
            border-radius: 4px;
            transition: all 0.2s;
        }
        .btn:hover {
            background-color: var(--btn-hover-bg);
            color: var(--btn-hover-color);
        }

        .sheet-status-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 15px;
            margin-top: 10px;
        }
        .sheet-card {
            border: 1px dashed var(--border-color);
            padding: 12px;
            background: var(--header-bg);
            border-radius: 4px;
        }
        .sheet-badge {
            display: inline-block;
            padding: 3px 8px;
            margin: 3px 3px 3px 0;
            border-radius: 3px;
            font-size: 12px;
            font-weight: bold;
            cursor: pointer;
            transition: transform 0.1s;
        }
        .sheet-badge:hover { transform: scale(1.05); }
        .badge-matched { background-color: #003300; color: #00ff00; border: 1px solid #00ff00; }
        .badge-old { background-color: #331100; color: #ffaa00; border: 1px solid #ffaa00; }
        .badge-new { background-color: #001133; color: #00ccff; border: 1px solid #00ccff; }

        [data-theme="traditional"] .badge-matched { background-color: #d4edda; color: #155724; border-color: #c3e6cb; }
        [data-theme="traditional"] .badge-old { background-color: #fff3cd; color: #856404; border-color: #ffeeba; }
        [data-theme="traditional"] .badge-new { background-color: #cce5ff; color: #004085; border-color: #b8daff; }

        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
            background-color: var(--bg-color);
            border: 1px solid var(--border-color);
            table-layout: fixed;
        }
        th, td {
            border: 1px solid var(--border-color);
            padding: 8px 10px;
            text-align: left;
            vertical-align: top;
            font-size: 13px;
            color: var(--text-color);
            word-break: break-all;
            white-space: pre-wrap;
            overflow: hidden;
            position: relative;
        }
        th { 
            background-color: var(--header-bg); 
            color: var(--text-color);
            font-weight: bold;
            cursor: pointer;
            user-select: none;
        }
        th:hover {
            background-color: var(--border-color);
            color: var(--bg-color);
        }

        .resizer {
            position: absolute;
            right: 0;
            top: 0;
            bottom: 0;
            width: 6px;
            cursor: col-resize;
            user-select: none;
            z-index: 10;
        }
        .resizer:hover, th.resizing .resizer, td.resizing .resizer {
            background-color: var(--link-color);
        }

        th:nth-child(1), td:nth-child(1) { width: 150px; }
        th:nth-child(2), td:nth-child(2) { width: 80px; }
        th:nth-child(3), td:nth-child(3) { width: 140px; }
        th:nth-child(4), td:nth-child(4) { width: 140px; }
        th:nth-child(5), td:nth-child(5) { width: 140px; }
        th:nth-child(6), td:nth-child(6) { width: 180px; }
        th:nth-child(7), td:nth-child(7) { width: 180px; }
        th:nth-child(8), td:nth-child(8) { width: 200px; }

        .diff-added { background-color: var(--added-bg); color: var(--added-color); padding: 1px 2px; border-radius: 2px; }
        .diff-deleted { background-color: var(--deleted-bg); color: var(--deleted-color); padding: 1px 2px; border-radius: 2px; }
        /* 專門畀 Formula Comparison 用嘅刪除線樣式 */
        .formula-deleted { background-color: var(--deleted-bg); color: var(--deleted-color); text-decoration: line-through; padding: 1px 2px; border-radius: 2px; }
        
        .header-row {
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 10px;
        }
        
        .pagination-bar {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-top: 10px;
            padding: 8px 12px;
            background: var(--header-bg);
            border: 1px solid var(--border-color);
            border-radius: 4px;
        }
    </style>
</head>
<body data-theme="hacker">

    <h1>Excel Difference History Report</h1>
    
    <div class="file-info">
        <div>
            <strong>Old File:</strong> <a href="__OLDURL__" target="_blank" class="file-link">__OLDFILEPATH__</a>
            <div class="meta-text">Last Modified By: __OLD_AUTHOR__ &nbsp;|&nbsp; Last Modified: __OLD_MOD__</div>
        </div>
        <div>
            <strong>New File:</strong> <a href="__NEWURL__" target="_blank" class="file-link">__NEWFILEPATH__</a>
            <div class="meta-text">Last Modified By: __NEW_AUTHOR__ &nbsp;|&nbsp; Last Modified: __NEW_MOD__</div>
        </div>
    </div>

    <div class="section-box" style="border-left: 4px solid var(--border-color);">
        <h3>Worksheet Status Overview</h3>
        <p class="meta-text" style="margin-bottom: 8px;">Clicking any sheet badge below will filter the view directly. Unpaired sheets (Old Only / New Only) require the checkbox below to be enabled.</p>
        
        <div class="sheet-status-grid" id="sheetStatusContainer"></div>
    </div>

    <div class="file-info">
        <div style="border-top: 1px dashed var(--border-color); padding-top: 8px;" class="header-row">
            <div>
                <strong>Generated At:</strong> <span>__TIMESTAMP__</span><br>
                <div id="summary" style="margin-top: 4px;"><strong>Total Changes:</strong> Loading...</div>
            </div>
            <div style="display: flex; align-items: center; gap: 12px; flex-wrap: wrap;">
                __UNCHANGED_CHECKBOX_HTML__
                <div>
                    <label for="sheetFilter" style="font-weight: bold; margin-right: 5px;">Sheet Filter:</label>
                    <select id="sheetFilter" class="btn" onchange="onFilterChange()">
                        <option value="">All Sheets</option>
                    </select>
                </div>
                <div>
                    <label for="pageSizeSelect" style="font-weight: bold; margin-right: 5px;">Per Page:</label>
                    <select id="pageSizeSelect" class="btn" onchange="onPageSizeChange()">
                        <option value="500">500</option>
                        <option value="1000">1000</option>
                        <option value="2500" selected>2500</option>
                        <option value="5000">5000</option>
                    </select>
                </div>
                <button class="btn" onclick="exportToCSV()">Export CSV</button>
                <div>
                    <button class="btn" onclick="setTheme('hacker')">Hacker</button>
                    <button class="btn" onclick="setTheme('traditional')">Traditional</button>
                </div>
            </div>
        </div>
    </div>

    <div id="main-container"></div>

    <script>
        const timelineData = __JSONDATA__;
        const sheetMeta = __SHEETMETA__;
        let currentSortCol = -1;
        let currentSortAsc = true;
        let currentPage = 1;

        function setTheme(themeName) {
            if (themeName === 'traditional') {
                document.body.setAttribute('data-theme', 'traditional');
            } else {
                document.body.setAttribute('data-theme', 'hacker');
            }
        }

        function renderSheetStatusOverview() {
            const container = document.getElementById('sheetStatusContainer');
            let html = '';

            html += `<div class="sheet-card">
                <strong>Matched [${sheetMeta.matched.length}]</strong>
                <div style="margin-top: 6px;">`;
            if (sheetMeta.matched.length > 0) {
                sheetMeta.matched.forEach(s => {
                    html += `<span class="sheet-badge badge-matched" onclick="selectSheetFromOverview('${s}')" title="Click to filter sheet">${escapeHtml(s)}</span>`;
                });
            } else {
                html += `<span class="meta-text" style="margin:0;">None</span>`;
            }
            html += `</div></div>`;

            html += `<div class="sheet-card" style="border-color: #ffaa00;">
                <strong style="color: #ffaa00;">Old Only [${sheetMeta.oldOnly.length}]</strong>
                <div style="margin-top: 6px;">`;
            if (sheetMeta.oldOnly.length > 0) {
                sheetMeta.oldOnly.forEach(s => {
                    html += `<span class="sheet-badge badge-old" onclick="selectSheetFromOverview('${s}')" title="Click to inspect deleted sheet">${escapeHtml(s)}</span>`;
                });
            } else {
                html += `<span class="meta-text" style="margin:0;">None</span>`;
            }
            html += `</div></div>`;

            html += `<div class="sheet-card" style="border-color: #00ccff;">
                <strong style="color: #00ccff;">New Only [${sheetMeta.newOnly.length}]</strong>
                <div style="margin-top: 6px;">`;
            if (sheetMeta.newOnly.length > 0) {
                sheetMeta.newOnly.forEach(s => {
                    html += `<span class="sheet-badge badge-new" onclick="selectSheetFromOverview('${s}')" title="Click to inspect added sheet">${escapeHtml(s)}</span>`;
                });
            } else {
                html += `<span class="meta-text" style="margin:0;">None</span>`;
            }
            html += `</div></div>`;

            container.innerHTML = html;
        }

        function selectSheetFromOverview(sheetName) {
            const selectEl = document.getElementById('sheetFilter');
            selectEl.value = sheetName;
            onFilterChange();
            window.scrollTo({ top: document.getElementById('main-container').offsetTop - 50, behavior: 'smooth' });
        }

        function computeDiffHtml(oldStr, newStr, isOldOnly, isNewOnly) {
            if (isOldOnly || isNewOnly) return '';
            if (!oldStr && !newStr) return '';
            if (oldStr === newStr) return '';
            if (!oldStr) return `<span class="diff-added">${escapeHtml(newStr)}</span>`;
            if (!newStr) return `<span class="formula-deleted">${escapeHtml(oldStr)}</span>`;
            return formulaDiff(oldStr, newStr);
        }

        function formulaDiff(o, n) {
            let minLen = Math.min(o.length, n.length);
            let commonPrefixLen = 0;
            while(commonPrefixLen < minLen && o[commonPrefixLen] === n[commonPrefixLen]) commonPrefixLen++;
            
            let commonSuffixLen = 0;
            while(commonSuffixLen < (minLen - commonPrefixLen) && 
                  o[o.length - 1 - commonSuffixLen] === n[n.length - 1 - commonSuffixLen]) commonSuffixLen++;

            let prefix = o.substring(0, commonPrefixLen);
            let oMid = o.substring(commonPrefixLen, o.length - commonSuffixLen);
            let nMid = n.substring(commonPrefixLen, n.length - commonSuffixLen);
            let suffix = o.substring(o.length - commonSuffixLen);

            let result = escapeHtml(prefix);
            if (oMid) result += `<span class="formula-deleted">${escapeHtml(oMid)}</span>`;
            if (nMid) result += `<span class="diff-added">${escapeHtml(nMid)}</span>`;
            result += escapeHtml(suffix);
            return result;
        }

        function escapeHtml(text) {
            return String(text).replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;");
        }

        function getAllChanges() {
            let changes = [];
            if (timelineData && timelineData.length > 0 && timelineData[0].events) {
                timelineData[0].events.forEach(ev => { changes = changes.concat(ev.changes); });
            }
            return changes;
        }

        function initSheetFilter() {
            const selectEl = document.getElementById('sheetFilter');
            let allSheets = [...sheetMeta.matched, ...sheetMeta.oldOnly, ...sheetMeta.newOnly];
            allSheets.sort();
            allSheets.forEach(sheet => {
                let opt = document.createElement('option');
                opt.value = sheet;
                opt.textContent = sheet;
                selectEl.appendChild(opt);
            });
        }

        function updateSummary(filteredCount, totalCount) {
            document.getElementById('summary').innerHTML = `<strong>Displayed Count:</strong> <b>${filteredCount}</b> / Total ${totalCount} records`;
        }

        function sortTable(colIdx) {
            if (currentSortCol === colIdx) {
                currentSortAsc = !currentSortAsc;
            } else {
                currentSortCol = colIdx;
                currentSortAsc = true;
            }
            currentPage = 1;
            filterAndRenderTable();
        }

        function onFilterChange() {
            currentPage = 1;
            filterAndRenderTable();
        }

        function onPageSizeChange() {
            currentPage = 1;
            filterAndRenderTable();
        }

        function changePage(targetPage) {
            currentPage = targetPage;
            filterAndRenderTable();
            window.scrollTo({ top: document.getElementById('main-container').offsetTop - 50, behavior: 'smooth' });
        }

        function parseCellAddress(addr) {
            let match = addr.match(/^([A-Z]+)(\d+)$/);
            if (!match) return { colNum: 0, rowNum: 0 };
            let colStr = match[1];
            let rowNum = parseInt(match[2], 10);
            let colVal = 0;
            for (let i = 0; i < colStr.length; i++) {
                colVal = colVal * 26 + (colStr.charCodeAt(i) - 64);
            }
            return { colNum: colVal, rowNum: rowNum };
        }

        let cachedFilteredChanges = null;

        function filterAndRenderTable() {
            const selectedSheet = document.getElementById('sheetFilter').value;
            const showUnchangedCheckbox = document.getElementById('showUnchanged');
            const showUnchanged = showUnchangedCheckbox ? showUnchangedCheckbox.checked : false;
            let changes = getAllChanges();

            cachedFilteredChanges = changes.filter(c => {
                if (selectedSheet && c.sheet !== selectedSheet) return false;
                
                if (!showUnchanged) {
                    if (!c.hasDiff || c.isOldOnly || c.isNewOnly) {
                        return false;
                    }
                }
                return true;
            });

            updateSummary(cachedFilteredChanges.length, changes.length);

            if (currentSortCol !== -1) {
                cachedFilteredChanges.sort((a, b) => {
                    let valA, valB;
                    if (currentSortCol === 0) { valA = a.sheet; valB = b.sheet; }
                    else if (currentSortCol === 1) {
                        let parsedA = parseCellAddress(a.address);
                        let parsedB = parseCellAddress(b.address);
                        if (parsedA.rowNum !== parsedB.rowNum) { valA = parsedA.rowNum; valB = parsedB.rowNum; }
                        else { valA = parsedA.colNum; valB = parsedB.colNum; }
                    }
                    else if (currentSortCol === 2) { valA = a.oldVal; valB = b.oldVal; }
                    else if (currentSortCol === 3) { valA = a.newVal; valB = b.newVal; }
                    else if (currentSortCol === 4) { valA = a.valDiff; valB = b.valDiff; }
                    else if (currentSortCol === 5) { valA = a.oldFormula; valB = b.oldFormula; }
                    else if (currentSortCol === 6) { valA = a.newFormula; valB = b.newFormula; }
                    else { valA = ''; valB = ''; }

                    if (typeof valA === 'number' && typeof valB === 'number') {
                        return currentSortAsc ? valA - valB : valB - valA;
                    }
                    let strA = String(valA).toLowerCase();
                    let strB = String(valB).toLowerCase();
                    if (strA < strB) return currentSortAsc ? -1 : 1;
                    if (strA > strB) return currentSortAsc ? 1 : -1;
                    return 0;
                });
            }

            const pageSize = parseInt(document.getElementById('pageSizeSelect').value, 10);
            const totalPages = Math.ceil(cachedFilteredChanges.length / pageSize) || 1;
            if (currentPage > totalPages) currentPage = totalPages;
            if (currentPage < 1) currentPage = 1;

            const startIndex = (currentPage - 1) * pageSize;
            const endIndex = startIndex + pageSize;
            const pageData = cachedFilteredChanges.slice(startIndex, endIndex);

            const container = document.getElementById('main-container');
            if (cachedFilteredChanges.length === 0) {
                container.innerHTML = '<div class="section-box"><p>No records matching the criteria. (Make sure "Show Unchanged & Unpaired" is enabled to view unchanged or unpaired sheets)</p></div>';
                return;
            }

            let html = '<div class="section-box"><table id="diffTable"><thead><tr>';
            const headers = ["Sheet", "Address", "Old Value", "New Value", "Difference", "Old Formula", "New Formula", "Formula Comparison"];
            headers.forEach((h, idx) => {
                let sortIndicator = (currentSortCol === idx) ? (currentSortAsc ? " ▲" : " ▼") : "";
                html += `<th onclick="handleHeaderClick(event, ${idx})">${h}${sortIndicator}<div class="resizer" onmousedown="initResize(event, ${idx})"></div></th>`;
            });
            html += `</tr></thead><tbody>`;
            
            pageData.forEach(c => {
                let formulaDiffHtml = computeDiffHtml(c.oldFormula, c.newFormula, c.isOldOnly, c.isNewOnly);
                
                let oldValHtml = escapeHtml(c.oldVal);
                let newValHtml = escapeHtml(c.newVal);
                let diffValHtml = c.valDiff;
                let oldFormulaHtml = escapeHtml(c.oldFormula);
                let newFormulaHtml = escapeHtml(c.newFormula);

                if (c.isOldOnly) {
                    oldValHtml = escapeHtml(c.oldVal);
                    newValHtml = '';
                    diffValHtml = '';
                    oldFormulaHtml = escapeHtml(c.oldFormula);
                    newFormulaHtml = '';
                } else if (c.isNewOnly) {
                    oldValHtml = '';
                    newValHtml = escapeHtml(c.newVal);
                    diffValHtml = '';
                    oldFormulaHtml = '';
                    newFormulaHtml = escapeHtml(c.newFormula);
                } else if (c.hasDiff) {
                    newValHtml = `<span class="diff-added">${escapeHtml(c.newVal)}</span>`;
                }

                html += `<tr>`;
                html += `<td>${escapeHtml(c.sheet)}<div class="resizer" onmousedown="initResize(event, 0)"></div></td>`;
                html += `<td>${escapeHtml(c.address)}<div class="resizer" onmousedown="initResize(event, 1)"></div></td>`;
                html += `<td>${oldValHtml}<div class="resizer" onmousedown="initResize(event, 2)"></div></td>`;
                html += `<td>${newValHtml}<div class="resizer" onmousedown="initResize(event, 3)"></div></td>`;
                html += `<td>${diffValHtml}<div class="resizer" onmousedown="initResize(event, 4)"></div></td>`;
                html += `<td>${oldFormulaHtml}<div class="resizer" onmousedown="initResize(event, 5)"></div></td>`;
                html += `<td>${newFormulaHtml}<div class="resizer" onmousedown="initResize(event, 6)"></div></td>`;
                html += `<td>${formulaDiffHtml}<div class="resizer" onmousedown="initResize(event, 7)"></div></td>`;
                html += `</tr>`;
            });
            html += `</tbody></table>`;

            // Pagination Controls
            html += `<div class="pagination-bar">
                <div>Page <b>${currentPage}</b> of <b>${totalPages}</b> (Showing items ${startIndex + 1} - ${Math.min(endIndex, cachedFilteredChanges.length)})</div>
                <div style="display: flex; gap: 5px;">
                    <button class="btn" ${currentPage === 1 ? 'disabled style="opacity:0.5; cursor:not-allowed;"' : 'onclick="changePage(1)"'}>First</button>
                    <button class="btn" ${currentPage === 1 ? 'disabled style="opacity:0.5; cursor:not-allowed;"' : 'onclick="changePage(' + (currentPage - 1) + ')"'}>Prev</button>
                    <button class="btn" ${currentPage === totalPages ? 'disabled style="opacity:0.5; cursor:not-allowed;"' : 'onclick="changePage(' + (currentPage + 1) + ')"'}>Next</button>
                    <button class="btn" ${currentPage === totalPages ? 'disabled style="opacity:0.5; cursor:not-allowed;"' : 'onclick="changePage(' + totalPages + ')"'}>Last</button>
                </div>
            </div>`;

            html += `</div>`;
            container.innerHTML = html;
        }

        let isResizing = false;
        let targetTh = null;
        let startX = 0;
        let startWidth = 0;

        function initResize(e, colIdx) {
            e.stopPropagation();
            isResizing = true;
            
            const table = document.getElementById('diffTable');
            targetTh = table.querySelectorAll('th')[colIdx];
            
            startX = e.clientX;
            startWidth = targetTh.offsetWidth;
            targetTh.classList.add('resizing');

            document.addEventListener('mousemove', doResize);
            document.addEventListener('mouseup', stopResize);
        }

        function doResize(e) {
            if (!isResizing || !targetTh) return;
            let widthDiff = e.clientX - startX;
            let newWidth = Math.max(50, startWidth + widthDiff);
            targetTh.style.width = newWidth + 'px';
        }

        function stopResize() {
            if (isResizing && targetTh) {
                targetTh.classList.remove('resizing');
            }
            isResizing = false;
            targetTh = null;
            document.removeEventListener('mousemove', doResize);
            document.removeEventListener('mouseup', stopResize);
        }

        function handleHeaderClick(event, idx) {
            if (event.target.classList.contains('resizer')) return;
            sortTable(idx);
        }

        function exportToCSV() {
            let changes = cachedFilteredChanges || getAllChanges();

            let csvContent = "\uFEFFSheet,Address,Old Value,New Value,Difference,Old Formula,New Formula\r\n";
            changes.forEach(c => {
                let cleanDiff = c.valDiff ? c.valDiff.replace(/<\/?[^>]+(>|$)/g, "") : "";
                let row = [
                    c.sheet, c.address,
                    `"${c.oldVal.replace(/"/g, '""')}"`,
                    `"${c.newVal.replace(/"/g, '""')}"`,
                    `"${cleanDiff.replace(/"/g, '""')}"`,
                    `"${c.oldFormula.replace(/"/g, '""')}"`,
                    `"${c.newFormula.replace(/"/g, '""')}"`
                ];
                csvContent += row.join(",") + "\r\n";
            });
            
            const blob = new Blob([csvContent], { type: 'text/csv;charset=utf-8;' });
            const link = document.createElement("a");
            link.href = URL.createObjectURL(blob);
            link.setAttribute("download", "excel_sheet_compare_export.csv");
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
        }

        window.onload = function() {
            renderSheetStatusOverview();
            initSheetFilter();
            filterAndRenderTable();
        };
    </script>
</body>
</html>"""

    html_content = html_template.replace('__OLDFILEPATH__', old_file_path) \
                                .replace('__NEWFILEPATH__', new_file_path) \
                                .replace('__OLDURL__', old_url) \
                                .replace('__NEWURL__', new_url) \
                                .replace('__OLD_AUTHOR__', old_author) \
                                .replace('__OLD_MOD__', old_mod) \
                                .replace('__NEW_AUTHOR__', new_author) \
                                .replace('__NEW_MOD__', new_mod) \
                                .replace('__TIMESTAMP__', timestamp) \
                                .replace('__JSONDATA__', json_data) \
                                .replace('__SHEETMETA__', sheet_meta_json) \
                                .replace('__UNCHANGED_CHECKBOX_HTML__', unchanged_checkbox_html)
    return html_content

if __name__ == "__main__":
    old_file_path = r"X:\MD7\Chain\2026Q1\Preliminary\2026.05.29\Chain VA for Section H (2026Q1 Preliminary) 2026.05.29.xlsm"
    new_file_path = r"X:\MD7\Chain\2026Q2\Preliminary\2026.08.31\Chain VA for Section H (2026Q2 Preliminary) 2026.08.31.xlsm"
    
    DEFAULT_LOG_FOLDER = r"D:\Pzone\excel_compare"
    
    old_data = excel_to_dict(old_file_path)
    new_data = excel_to_dict(new_file_path)
    
    report_path = generate_diff_report(
        old_data, 
        new_data, 
        old_file_path, 
        new_file_path,
        output_dir=os.path.join(DEFAULT_LOG_FOLDER, 'diff_reports'),
        include_unchanged_cells=True
    )
    print(f"Test successful! Please open the HTML report:\n{report_path}")
